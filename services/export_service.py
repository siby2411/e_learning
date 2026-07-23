#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import csv
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from models.database import db

class ExportService:
    @staticmethod
    def export_to_csv(data, headers, filename):
        filepath = os.path.join('exports', f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        os.makedirs('exports', exist_ok=True)
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)
        return filepath

    @staticmethod
    def export_to_excel(data, headers, filename):
        filepath = os.path.join('exports', f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        os.makedirs('exports', exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Données"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, row in enumerate(data, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filepath)
        return filepath

    @staticmethod
    def export_enrollments(format='csv'):
        rows = db.fetch_all("""
            SELECT 
                u.email, u.first_name, u.last_name,
                c.title as course_title,
                e.status, e.payment_status, e.progress_percentage,
                e.start_date, e.completion_date
            FROM enrollments e
            JOIN users u ON e.user_id = u.id
            JOIN courses c ON e.course_id = c.id
            ORDER BY e.id
        """)
        headers = ['Email', 'Prénom', 'Nom', 'Cours', 'Statut', 'Paiement', 'Progression (%)', 'Date début', 'Date fin']
        data = [[
            row['email'], row['first_name'], row['last_name'],
            row['course_title'], row['status'], row['payment_status'],
            row['progress_percentage'], row['start_date'], row['completion_date']
        ] for row in rows]
        
        if format == 'csv':
            return ExportService.export_to_csv(data, headers, 'enrollments')
        else:
            return ExportService.export_to_excel(data, headers, 'enrollments')

    @staticmethod
    def export_grades(format='csv'):
        rows = db.fetch_all("""
            SELECT 
                u.email, u.first_name, u.last_name,
                c.title as course_title,
                ex.title as exam_title,
                g.score, g.is_passed, g.graded_at
            FROM grades g
            JOIN enrollments e ON g.enrollment_id = e.id
            JOIN users u ON e.user_id = u.id
            JOIN courses c ON e.course_id = c.id
            JOIN exams ex ON g.exam_id = ex.id
            ORDER BY g.id
        """)
        headers = ['Email', 'Prénom', 'Nom', 'Cours', 'Examen', 'Note', 'Réussi', 'Date notation']
        data = [[
            row['email'], row['first_name'], row['last_name'],
            row['course_title'], row['exam_title'],
            row['score'], 'Oui' if row['is_passed'] else 'Non',
            row['graded_at']
        ] for row in rows]
        
        if format == 'csv':
            return ExportService.export_to_csv(data, headers, 'grades')
        else:
            return ExportService.export_to_excel(data, headers, 'grades')
